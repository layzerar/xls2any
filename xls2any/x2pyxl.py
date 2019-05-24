# -*- coding: utf-8 -*-

import re
import os
import sys
import ctypes
import datetime
import functools
import itertools
import collections
from decimal import Decimal

import openpyxl

from . import utils

Ctx = utils.Ctx


RANGE_SEP = ':'
RANGE_NIL = '...'
COLKEY_TOKEN = '@'
COLKEY_REGEX = \
    re.compile(r'^(@[^,$:]+)$', re.IGNORECASE)
CELLX1_REGEX = \
    re.compile(r'^([A-Z]+)([1-9][0-9]*)$', re.IGNORECASE)
CELLX2_REGEX = \
    re.compile(
        r'^(?:(@[^,$:]+)([,$][1-9][0-9]*)|([A-Z]+)([1-9][0-9]*))$',
        re.IGNORECASE)
COLUMN_REGEX = \
    re.compile(r'^([A-Z]+)$', re.IGNORECASE)
VINDEX_REGEX = \
    re.compile(r'^([1-9][0-9]*)$', re.IGNORECASE)
RANGE1_REGEX = \
    re.compile(
        r'^(?:(@[^,$:]+)([,$][1-9][0-9]*)?|([A-Z]+)?([1-9][0-9]*)?)'
        r':(?:(@[^,$:]+)([,$][1-9][0-9]*)?|([A-Z]+)?([1-9][0-9]*)?)$',
        re.IGNORECASE)
RANGE2_REGEX = \
    re.compile(
        r'^(?:(@[^,$:]+)|([A-Z]+))$',
        re.IGNORECASE)
RANGE3_REGEX = VINDEX_REGEX

DEFAULT_DATETIME = datetime.datetime(1900, 1, 1)


def build_column(col):
    digits = []
    while col > 0:
        digits.append(chr((col - 1) % 26 + 65))
        col //= 26
    return ''.join(reversed(digits))


def parse_column(expr):
    if isinstance(expr, str):
        matches = COLUMN_REGEX.match(expr.upper())
        if matches:
            col = 0
            for ich in matches.group(1):
                col = col * 26 + ord(ich) - 64
            return col
    Ctx.throw('错误的列号格式：{0!r}', expr)


RangeArgs = collections.namedtuple(
    'RangeArgs',
    ['hoff', 'voff', 'hnum', 'vnum'],
)


def cell_args(lcol, lrow):
    return RangeArgs(lcol - 1, lrow - 1, 1, 1)


def cell_expr(lcol, lrow):
    return build_column(lcol) + str(lrow)


def range_args(lcol, lrow, hcol, hrow):
    return RangeArgs(lcol - 1, lrow - 1, hcol - lcol + 1, hrow - lrow + 1)


def range_expr(lcol, lrow, hcol, hrow):
    return RANGE_SEP.join([
        build_column(lcol) + str(lrow),
        build_column(hcol) + str(hrow),
    ])


def parse_cell(expr, headers=None):
    if not isinstance(expr, str):
        Ctx.throw('错误的单元格式：{0!r}', expr)
    matches = CELLX2_REGEX.match(expr.strip())
    if not matches:
        Ctx.throw('错误的单元格式：{0!r}', expr)
    if matches.group(1):
        lkey = matches.group(1)
        lcol = headers.get(lkey[1:], [0])[0] if headers else 0
        if lcol == 0:
            Ctx.throw('无法定位指定列：{0!r}', lkey)
        lrow = int(matches.group(2)[1:])
    else:
        lcol = parse_column(matches.group(3))
        lrow = int(matches.group(4))
    return (
        cell_expr(lcol, lrow),
        cell_args(lcol, lrow),
    )


def parse_range(expr, max_col, max_row, headers=None, min_col=1, min_row=1):
    def make_return(lcol, lrow, hcol, hrow):
        return (
            range_expr(lcol, lrow, hcol, hrow),
            range_args(lcol, lrow, hcol, hrow),
        )

    def make_range1(matches):
        if matches.group(1):
            lkey = matches.group(1)
            lcol = headers.get(lkey[1:], [0])[0] if headers else 0
            if lcol == 0:
                Ctx.throw('无法定位指定列：{0!r}', lkey)
            lrow = int(matches.group(2)[1:]) if matches.group(2) else min_row
        else:
            lcol = parse_column(matches.group(3)) \
                if matches.group(3) else min_col
            lrow = int(matches.group(4)) if matches.group(4) else min_row
        if matches.group(5):
            hkey = matches.group(5)
            hcol = headers.get(hkey[1:], [0])[0] if headers else 0
            if hcol == 0:
                Ctx.throw('无法定位指定列：{0!r}', hkey)
            hrow = int(matches.group(6)[1:]) if matches.group(6) else max_row
        else:
            hcol = parse_column(matches.group(7)) \
                if matches.group(7) else max_col
            hrow = int(matches.group(8)) if matches.group(8) else max_row
        return make_return(lcol, lrow, hcol, hrow)

    def make_range2(matches):
        if matches.group(1):
            ikey = matches.group(1)
            icol = headers.get(ikey[1:], [0])[0] if headers else 0
            if icol == 0:
                Ctx.throw('无法定位指定列：{0!r}', ikey)
        else:
            icol = parse_column(matches.group(2))
        return make_return(icol, min_row, icol, max_row)

    def make_range3(matches):
        irow = int(matches.group(1))
        return make_return(min_col, irow, max_col, irow)

    if isinstance(expr, str):
        if expr.strip() == RANGE_SEP:
            return make_return(min_col, min_row, max_col, max_row)

        for regex, make_range in [(RANGE1_REGEX, make_range1),
                                  (RANGE2_REGEX, make_range2),
                                  (RANGE3_REGEX, make_range3)]:
            matches = regex.match(expr.strip())
            if matches:
                return make_range(matches)

    Ctx.throw('错误的区域格式：{0!r}', expr)


try:
    _libc_strcmp = ctypes.cdll.msvcrt.strcmp
except OSError:
    try:
        _libc_strcmp = ctypes.CDLL("libc.so.6").strcmp
    except OSError:
        def _libc_strcmp(str1, str2):
            if str1 == str2:
                return 0
            else:
                return -1 if str1 < str2 else 1


TOSTRING_TYPES = {
    int:                 str,
    float:               str,
    Decimal:             str,
    bool:                lambda v1: 'TRUE' if v1 else 'FALSE',
    datetime.date:       str,
    datetime.datetime:   str,
    datetime.time:       str,
    str:                 lambda v1: v1.strip(),
    type(None):          lambda v1: '',
}


def xstr(val1):
    conv = TOSTRING_TYPES.get(type(val1))
    if conv is None:
        Ctx.throw('不能将该类型转化为文本：{0}', tp_val1.__name__)
    return conv(val1)


def _cmp_num_num(val1, val2):
    if val1 == val2:
        return 0
    elif val1 < val2:
        return -1
    else:
        return 1


def _cmp_str_str(val1, val2):
    val1 = val1.strip()
    val2 = val2.strip()
    diff = len(val1) - len(val2)
    if diff == 0:
        return _libc_strcmp(val1, val2)
    elif diff < 0:
        return -1
    else:
        return 1


def _cmp_str_num(val1, val2):
    try:
        fit1 = float(val1.strip())
    except ValueError:
        return _cmp_str_str(val1, str(val2))
    else:
        return _cmp_num_num(fit1, val2)


def _cmp_str_bool(val1, val2):
    try:
        fit1 = float(val1.strip())
    except ValueError:
        return _cmp_str_str(val1, 'TRUE' if val2 else 'FALSE')
    else:
        return _cmp_num_num(fit1, val2)


def _cmp_str_date(val1, val2):
    return _cmp_str_str(val1, str(val2))


def _cmp_date_datetime(val1, val2):
    fit1 = datetime.datetime.combine(val1, datetime.time())
    return _cmp_num_num(fit1, val2)


def _cmp_datetime_date(val1, val2):
    fit2 = datetime.datetime.combine(val2, datetime.time())
    return _cmp_num_num(val1, fit2)


XCMP_ALL_TYPES = {
    int: 10,
    float: 11,
    Decimal: 12,
    bool: 13,
    datetime.date: 21,
    datetime.datetime: 22,
    datetime.time: 23,
    str: 30,
    tuple: 40,
    type(None): 50,
}
XCMP_FIT_TYPES = {
    # special types
    (bool, bool):               _cmp_num_num,
    (type(None), type(None)):   (lambda v1, v2: 0),

    # numeric types
    (int, int):                 _cmp_num_num,
    (int, float):               _cmp_num_num,
    (int, Decimal):             _cmp_num_num,
    (float, int):               _cmp_num_num,
    (float, float):             _cmp_num_num,
    (float, Decimal):           _cmp_num_num,
    (Decimal, int):             _cmp_num_num,
    (Decimal, float):           _cmp_num_num,
    (Decimal, Decimal):         _cmp_num_num,

    # downcast types
    (str, int):                 _cmp_str_num,
    (str, float):               _cmp_str_num,
    (str, Decimal):             _cmp_str_num,
    (str, bool):                _cmp_str_bool,
    (str, datetime.date):       _cmp_str_date,
    (str, datetime.time):       _cmp_str_date,
    (str, datetime.datetime):   _cmp_str_date,
    (str, str):                 _cmp_str_str,
    (str, type(None)):          lambda v1, _: _cmp_str_str(v1, ''),
    (type(None), str):          lambda _, v2: _cmp_str_str('', v2),

    # datetime types
    (datetime.date, datetime.date):         _cmp_num_num,
    (datetime.date, datetime.datetime):     _cmp_date_datetime,
    (datetime.time, datetime.time):         _cmp_num_num,
    (datetime.datetime, datetime.date):     _cmp_datetime_date,
    (datetime.datetime, datetime.datetime): _cmp_num_num,
}


def xeq_(val1, val2):
    return xcmp_(val1, val2) == 0


def xlt_(val1, val2):
    return xcmp_(val1, val2) < 0


def xle_(val1, val2):
    return xcmp_(val1, val2) <= 0


def xgt_(val1, val2):
    return xcmp_(val1, val2) > 0


def xge_(val1, val2):
    return xcmp_(val1, val2) >= 0


def xcmp_(val1, val2):
    if val1 is val2:
        return 0
    tp_val1 = type(val1)
    tp_val2 = type(val2)
    if tp_val1 is tuple and tp_val2 is tuple:
        diff = len(val1) - len(val2)
        if diff == 0:
            for res in itertools.starmap(xcmp_, zip(val1, val2)):
                if res != 0:
                    return res
            return 0
        elif diff < 0:
            return -1
        else:
            return 1
    else:
        _cmp_impl = XCMP_FIT_TYPES.get((tp_val1, tp_val2))
        if _cmp_impl is not None:
            return _cmp_impl(val1, val2)
        _cmp_impl = XCMP_FIT_TYPES.get((tp_val2, tp_val1))
        if _cmp_impl is not None:
            return _cmp_impl(val2, val1)
        if tp_val1 not in XCMP_ALL_TYPES or tp_val2 not in XCMP_ALL_TYPES:
            Ctx.throw('不支持该类型之间的比较：{0} <-> {1}', tp_val1.__name__, tp_val2.__name__)
        return _cmp_num_num(XCMP_ALL_TYPES[tp_val1], XCMP_ALL_TYPES[tp_val2])


class XlRowView(object):

    def __init__(self, sheet, array, offset, vindex):
        self._sheet = sheet
        self._array = array
        self._offset = offset
        self._vindex = vindex

    def __len__(self):
        return len(self._array)

    def __getitem__(self, key):
        if isinstance(key, slice):
            if key.step is not None:
                Ctx.throw('行对象不支持间隔切片：{0!r}', key)
            if key.start is None and key.stop is None:
                return self
            hidx1 = 1 if key.start is None else key.start
            hidx2 = len(self._array) + 1 if key.stop is None else key.stop
            hidx1 = self.hidx(hidx1) - self._offset
            hidx2 = self.hidx(hidx2) - self._offset
            return self.cut(min(hidx1, hidx2), abs(hidx2 - hidx1) + 1)
        return self.valx(key)

    @property
    def vidx(self):
        return self._vindex

    def hidx(self, key, multi=False):
        if isinstance(key, int):
            if key <= 0:
                Ctx.throw('无法定位指定列：{0!r}', key)
            hoff = self._offset
            hmax = hoff + len(self._array)
            return self._sheet.hidx(hoff + key, multi=multi, hoff=hoff, hmax=hmax)
        else:
            hoff = self._offset
            hmax = hoff + len(self._array)
            return self._sheet.hidx(key, multi=multi, hoff=hoff, hmax=hmax)

    def expr(self, key):
        return build_column(self.hidx(key)) + str(self._vindex)

    def keys(self, *idxs, token=True):
        def impl(idxs, offset, maxidx):
            for hidx in idxs:
                if not isinstance(hidx, int) or hidx <= 0:
                    Ctx.throw('索引必须是大于零的整数：{0!r}', hidx)
                if hidx > maxidx:
                    Ctx.throw('索引超出区域范围：{0!r}', hidx)
                yield offset + hidx
        idxs = tuple(impl(idxs, self._offset, len(self._array)))
        return self._sheet.keys(*idxs, token=token)

    def vals(self, *keys):
        if not keys:
            return tuple(elm.value for elm in self._array)
        else:
            return tuple(self.valx(key) for key in keys)

    def valx(self, key):
        col = self.hidx(key) - self._offset
        if 0 < col <= len(self._array):
            return self._array[col - 1].value
        return None

    def cut(self, key, size=1):
        for elm in self.slc(key, size, 1):
            return elm
        return None

    def slc(self, key, size=1, num=0):
        if not isinstance(size, int) or size <= 0:
            Ctx.throw('分组大小必须是大于零的整数：{0!r}', size)
        if not isinstance(num, int) or num < 0:
            Ctx.throw('分组数量必须是不为负的整数：{0!r}', num)
        if num == 0:
            offsets = [col - self._offset - 1 for col in self.hidx(key, multi=True)]
        else:
            offset = self.hidx(key) - self._offset
            offsets = [idx * size + offset - 1 for idx in range(num)]
        if offsets[0] < 0 or offsets[-1] + size > len(self._array):
            Ctx.throw('分组超出区域范围：{0!r},{1},{2}', key, size, num)
        for offset in offsets:
            yield type(self)(
                self._sheet,
                self._array[offset:offset+size],
                self._offset + offset,
                self._vindex,
            )

    def aslist(self, skip_none=False):
        if not skip_none:
            return [elm.value for elm in self._array]
        else:
            return [elm.value for elm in self._array if not xeq_(elm.value, '')]

    def asdict(self, *keys, skip_none=False):
        if not keys:
            idx0 = self._offset + 1
            idxs = tuple(range(idx0, idx0 + len(self._array)))
            keys = self._sheet.keys(*idxs, token=False)
        if not skip_none:
            return {str(key): elm.value for key, elm in zip(keys, self._array)}
        else:
            return {str(key): elm.value for key, elm in zip(keys, self._array) if not xeq_(elm.value, '')}


class SheetView(object):

    def __init__(self, filepath, sheetname, worksheet,
                 headers=None, beg_col=None, beg_row=None, end_col=None, end_row=None):
        self._filepath = filepath
        self._filename = os.path.basename(filepath)
        self._sheetname = sheetname
        self._worksheet = worksheet
        self._headers = headers or {}
        self._beg_col = beg_col or 1
        self._beg_row = beg_row or 1
        self._end_col = end_col or self._worksheet.max_column
        self._end_row = end_row or self._worksheet.max_row
        self._cur_row = None
        self._caches = {}

    def __str__(self):
        return '{0}#{1}'.format(self._filename, self._sheetname)

    def __iter__(self):
        return self[RANGE_SEP]

    def __getitem__(self, expr):
        if isinstance(expr, slice):
            if expr.step is not None:
                Ctx.throw('工作表不支持间隔切片：{0!r}', expr)
            expr = RANGE_SEP.join([
                str(expr.start or ''),
                str(expr.stop or ''),
            ])
        if expr == RANGE_NIL:
            return
        max_row = self._worksheet.max_row
        max_col = self._worksheet.max_column
        if max_row <= 0 or max_col <= 0:
            return
        slc_expr, args = parse_range(
            expr, max_col, max_row, self._headers, min_col=self._beg_col)
        for idx, row in xslice(self._worksheet[slc_expr], self._beg_row, self._end_row):
            Ctx.set_ctx(str(self), args.voff + idx)
            self._cur_row = XlRowView(self, row, args.hoff, args.voff + idx)
            yield self._cur_row

    @property
    def vidx(self):
        return self._cur_row.vidx if self._cur_row else 0

    def hidx(self, key, multi=False, hoff=0, hmax=sys.maxsize):
        if isinstance(key, str):
            if key.startswith(COLKEY_TOKEN):
                ret = [x for x in self._headers.get(key[1:], []) if hoff < x and x <= hmax]
                col = ret if ret else 0
            else:
                col = parse_column(key)
        elif isinstance(key, int):
            col = key if hoff < key and key <= hmax else 0
        else:
            col = 0
        if isinstance(col, int):
            if col == 0:
                Ctx.throw('无法定位指定列：{0!r}', key)
            return [col] if multi else col
        else:
            return col if multi else col[0]

    def keys(self, *idxs, token=True):
        def impl(idxs, headers):
            for hidx in idxs:
                if not isinstance(hidx, int) or hidx <= 0:
                    Ctx.throw('索引必须是大于零的整数：{0!r}', hidx)
                if hidx not in headers:
                    Ctx.throw('索引指向的列不包含表头：{0!r}', hidx)
                if token:
                    yield COLKEY_TOKEN + headers[hidx]
                else:
                    yield headers[hidx]
        return tuple(impl(idxs, self._headers))

    def rowx(self, vidx=None):
        if vidx is None:
            return self._cur_row
        if self._worksheet.max_row <= 0 \
                or self._worksheet.max_column <= 0:
            return None
        if isinstance(vidx, int):
            if vidx <= 0:
                Ctx.throw('无法定位指定行：{0!r}', vidx)
        else:
            if not isinstance(vidx, str) or \
                    not VINDEX_REGEX.match(vidx.strip()):
                Ctx.throw('无法定位指定行：{0!r}', vidx)
            vidx = int(vidx.strip())
        return XlRowView(self, self._worksheet[str(vidx)], 0, vidx)

    def valx(self, expr):
        cx1_expr, _ = parse_cell(expr, self._headers)
        return self._worksheet[cx1_expr].value

    def exprcol(self, expr, to_int=False):
        _, args = parse_cell(expr, self._headers)
        return args.hoff + 1 if to_int else build_column(args.hoff + 1)

    def exprrow(self, expr):
        _, args = parse_cell(expr, self._headers)
        return args.voff + 1

    def exproff(self, expr, hoff=1, voff=0):
        if not isinstance(hoff, int):
            Ctx.throw('水平位移量必须是整数：{0!r}', hoff)
        if not isinstance(voff, int):
            Ctx.throw('垂直位移量必须是整数：{0!r}', voff)
        cx1_expr, args = parse_cell(expr, self._headers)
        if hoff == 0 and voff == 0:
            return cx1_expr
        hidx = args.hoff + hoff + 1
        vidx = args.voff + voff + 1
        if hidx <= 0 or vidx <= 0:
            Ctx.throw('单元偏移超出范围：{0!r},{1},{2}', expr, hoff, voff)
        return cell_expr(hidx, vidx)

    def rehead(self, vidx, hbeg=None, vbeg=None, hend=None, vend=None):
        max_row = self._worksheet.max_row
        max_col = self._worksheet.max_column
        if max_row <= 0 or max_col <= 0:
            return self
        if isinstance(vidx, int):
            if vidx <= 0:
                Ctx.throw('无法定位指定行：{0!r}', vidx)
        else:
            if not isinstance(vidx, str) or \
                    not VINDEX_REGEX.match(vidx.strip()):
                Ctx.throw('无法定位指定行：{0!r}', vidx)
            vidx = int(vidx.strip())
        slc_expr, _ = parse_range(
            str(vidx), max_col, max_row, min_col=self._beg_col)
        for head_row in self._worksheet[slc_expr]:
            headers = {}
            for col, cell in enumerate(head_row, 1):
                if xeq_(cell.value, ''):
                    continue
                key = str(cell.value).strip()
                if not key:
                    continue
                vals = headers.get(key, [])
                headers[key] = vals + [col]
            for key, vals in list(headers.items()):
                headers.update((val, key) for val in vals)
            return type(self)(
                self._filepath,
                self._sheetname,
                self._worksheet,
                headers=headers,
                beg_col=hbeg,
                beg_row=vbeg,
                end_col=hend,
                end_row=vend,
            )
        return self

    @utils.cachedmethod('_caches', 'select:{1}', tuple, iter)
    def select(self, expr):
        max_row = self._worksheet.max_row
        max_col = self._worksheet.max_column
        if max_row <= 0 or max_col <= 0:
            return
        slc_expr, args = parse_range(
            expr, max_col, max_row, self._headers, min_col=self._beg_col)
        if slc_expr is not None:
            for idx, row in xslice(self._worksheet[slc_expr], self._beg_row, self._end_row):
                yield XlRowView(self, row, args.hoff, args.voff + idx)

    def locate(self, ltag, htag, loff=1, hoff=-1):
        hbeg, vbeg, hend, vend = None, None, None, None
        for row in self.select(RANGE_SEP):
            for idx, val in enumerate(row.vals(), 1):
                if val is None:
                    continue
                if vbeg is None and xeq_(ltag, val):
                    vbeg = row.vidx
                    hbeg = row.hidx(idx)
                    continue
                if vend is None and xeq_(htag, val):
                    vend = row.vidx
                    hend = row.hidx(idx)
                    break
        if vbeg is None:
            Ctx.error('找不匹配的起始标签：{0}', ltag)
            return self[RANGE_NIL]
        if vend is None:
            Ctx.error('找不匹配的结束标签：{0}', htag)
            return self[RANGE_NIL]
        return self.rehead(
            vbeg,
            hbeg=hbeg + 1,
            vbeg=vbeg + loff,
            hend=hend + 0,
            vend=vend + hoff)

    @utils.cachedmethod('_caches', '__index:{1}:{2}')
    def __index(self, tab, keys):
        index = {}
        for row in self.select(tab):
            ukey = tuple(xstr(x) for x in row.vals(*keys))
            vals = index.get(ukey)
            if vals is None:
                vals = index[ukey] = []
            vals.append(row)
        return {ukey: tuple(vals) for ukey, vals in index.items()}

    def findall(self, val1, tab, *keys):
        if isinstance(val1, (tuple, list)):
            if len(val1) <= 0:
                Ctx.throw('目标元素个数必须大于零：{0!r}', val1)
            if not keys:
                keys = tuple(range(1, len(val1) + 1))
            elif len(keys) != len(val1):
                Ctx.throw('目标元素个数和索引列数不一致：{0!r}', keys)
            ukey = tuple(xstr(x) for x in val1)
        else:
            if not keys:
                keys = (1,)
            elif len(keys) != len(val1):
                Ctx.throw('目标元素个数和索引列数不一致：{0!r}', keys)
            ukey = (xstr(val1),)
        return self.__index(tab, keys).get(ukey, ())

    def findone(self, val1, tab, *keys):
        for row in self.findall(val1, tab):
            return row
        Ctx.throw('指定区域\'{0}\'!{1}找不到对应值{2!r}', str(self), tab, val1)

    def vlookup(self, val1, tab, idx):
        return self.findone(val1, tab).valx(idx)

    def chkuniq(self, tab):
        max_row = self._worksheet.max_row
        max_col = self._worksheet.max_column
        if max_row <= 0 or max_col <= 0:
            return
        slc_expr, args = parse_range(
            tab, max_col, max_row, self._headers, min_col=self._beg_col)
        keys = tuple(range(1, args.hnum + 1))
        iterable = (
            XlRowView(self, row, args.hoff, args.voff + idx)
            for idx, row in xslice(self._worksheet[slc_expr], self._beg_row, self._end_row)
        )
        for key, group in xgroupby(iterable, *keys):
            first = None
            for row in group:
                if first is None:
                    first = row
                    continue
                Ctx.error('指定区域\'{0}\'!{1}含有重复值{2!r}：{3} <-> {4}',
                          str(self), tab, key, row.vidx, first.vidx)


def xslice(rows, vbeg=1, vend=sys.maxsize):
    if not hasattr(rows, '__iter__'):
        Ctx.throw('xslice 的传入参数必须是集合')
    for idx, row in enumerate(rows, 1):
        if idx > vend:
            break
        elif idx < vbeg:
            continue
        yield idx, row


def xrequire(rows, *keys, over=0):
    if not hasattr(rows, '__iter__'):
        Ctx.throw('xrequire 的传入参数必须是集合')
    blanks = 0
    for row in rows:
        if all(not xeq_(row.valx(key), '') for key in keys):
            blanks = 0
            yield row
        else:
            blanks += 1
            if 0 < over <= blanks:
                break


def xgroupby(rows, *keys, asc=True, required=True):
    if not hasattr(rows, '__iter__'):
        Ctx.throw('xgroupby 的传入参数必须是集合')

    def rowcmp(row1, row2):
        key1 = tuple(row1.valx(key) for key in keys)
        key2 = tuple(row2.valx(key) for key in keys)
        return xcmp_(key1, key2)

    origin_rows = rows if not required else xrequire(rows, *keys)
    sorted_rows = sorted(origin_rows, key=functools.cmp_to_key(rowcmp), reverse=not asc)
    for key, group in itertools.groupby(sorted_rows, key=functools.cmp_to_key(rowcmp)):
        yield key.obj, group


def load_worksheet(filepath, sheetname, head=0):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        ws = wb[sheetname]
    except IOError:
        Ctx.abort('无法打开目标工作簿：{0}', filepath)
    except KeyError:
        Ctx.abort('无法打开目标工作表：{0}#{1}', filepath, sheetname)

    sheet_view = SheetView(filepath, sheetname, ws)
    if 0 < head <= ws.max_row:
        sheet_view = sheet_view.rehead(head)
    return sheet_view
